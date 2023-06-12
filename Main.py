from openpyxl import Workbook
import openpyxl
import pathlib
import os

# fungsi clear terminal
def clear_terminal():
    if os.name == 'posix':
        os.system('clear')
    elif os.name == 'nt':
        os.system('cls')

# cek data base
def cek_data_base(file):
    if file.exists():
        pass
    else:
        file = Workbook()
        sheet = file.active
        sheet['A1'] = 'NIK'
        sheet['B1'] = 'Nama'
        sheet['C1'] = 'tempat_lahir'
        sheet['D1'] = 'tanggal_lahir'
        sheet['E1'] = 'bullan_lahir'
        sheet['F1'] = 'tahun_lahir'
        sheet['G1'] = 'alamat'
        sheet['H1'] = 'rt'
        sheet['I1'] = 'rw'
        sheet['J1'] = 'kelurahan'
        sheet['K1'] = 'kecamatan'
        sheet['L1'] = 'kabupater'
        sheet['M1'] = 'provinsi'
        sheet['N1'] = 'jenis_kelamin'
        sheet['O1'] = 'agama'


        file.save('Data_Base.xlsx')
        print('Data Base Dibuat!!!')

def data_cek(data):
    list_data = data
    for i in list_data:
        if i == '':
            return False
            break
    else:
        return True

def tambah_data():
    clear_terminal()
    nik = input('masukan nik\t\t:')
    nama = input('masukan nama\t\t:')
    tempat_lahir = input('masukan tempat lahir\t:')
    tanggal_lahir = input('masukan tanggal lahir\t:')
    bulan_lahir = input('masukan bulan lahir\t:')
    tahun_lahir = input('masukan tahun lahir\t:')
    alamat = input('masukan alamat\t\t:')
    rt = input('masukan rt\t\t:')
    rw = input('masukan rw\t\t:')
    kelurahan = input('masukan kelurahan\t:')
    kecamatan = input('masukan kecamatan\t:')
    kabupaten = input('masukan kabupaten\t:')
    provinsi = input('masukan provinsi\t:')
    jenis_kelamin = input('masukan jenis_kelamin\t:')
    agama = input('masukan agama\t\t:')

    list_data = [nik, nama.upper(), tempat_lahir.upper(), tanggal_lahir, bulan_lahir, tahun_lahir, alamat.upper(), rt, rw, kelurahan.upper(), kecamatan.upper(), kabupaten.upper(), provinsi.upper(), jenis_kelamin.upper(), agama.upper()]
    cek_data = data_cek(list_data)

    if cek_data == False:
        clear_terminal()
        print('error', 'Data belum lengkap')
        ulang = input('ulangi lagi(y/n):')
        if ulang == 'y':
            tambah_data()
        else:
            pass
    else:
        file = openpyxl.load_workbook('Data_Base.xlsx')
        sheet = file.active
        baris = sheet.max_row+1
        kolom = 0
        # entri data ke data base
        file = openpyxl.load_workbook('Data_Base.xlsx')
        sheet = file.active
        for i in list_data:
            kolom += 1
            sheet.cell(column=kolom, row=baris, value=i)
        file.save('Data_Base.xlsx')
        clear_terminal()
        print('info', 'data tersimpan!!!')

def lihat_data():
    file = openpyxl.load_workbook('Data_Base.xlsx')
    sheet = file.active
    

    while(True):
        clear_terminal()
        for i in range(2, sheet.max_row+1):
            nik = sheet.cell(column=1, row=i)
            nama = sheet.cell(column=2, row=i)
            print(i-1,f'{nik.value}', f'{nama.value}')
        baris = int(input('Masukan baris yang akan ditampilkan : '))
        lihat_ktp(sheet, baris+1)
        ulang = input('tampilkan ktp lagi?(y/n)')
        if ulang == 'y':
            pass
        else:
            break

    file.close()



def lihat_ktp(data, baris):
    sheet = data
    
    nik = sheet.cell(column= 1, row=baris)
    nama = sheet.cell(column= 2, row=baris)
    tempat_lahir = sheet.cell(column= 3, row=baris)
    tanggal_lahir = sheet.cell(column= 4, row=baris)
    bulan_lahir = sheet.cell(column= 5, row=baris)
    tahun_lahir = sheet.cell(column= 6, row=baris)
    alamat = sheet.cell(column= 7, row=baris)
    rt = sheet.cell(column= 8, row=baris)
    rw = sheet.cell(column= 9, row=baris)
    kelurahan = sheet.cell(column= 10, row=baris)
    kecamatan = sheet.cell(column= 11, row=baris)
    kabupaten = sheet.cell(column= 12, row=baris)
    provinsi = sheet.cell(column= 13, row=baris)
    jenis_kelamin = sheet.cell(column= 14, row=baris)
    agama = sheet.cell(column= 15, row=baris)

    clear_terminal()
    print('-'*80)
    print('|'+' '*(32-int(len(provinsi.value)/2))+ 'PROVINSI '+ provinsi.value+' '*(37+int(len(provinsi.value)/2)-len(provinsi.value))+'|')
    print('|'+' '*(31-int((len(kabupaten.value)/2)))+ 'KABUPATEN '+ kabupaten.value+' '*(37+int(len(kabupaten.value)/2)-len(kabupaten.value))+'|')
    print('|NIK\t  : '+nik.value+' '*(67-int(len(nik.value)))+'|')
    print('|'+' '*78+'|')
    print('|Nama\t\t  : '+nama.value+' '*(59-int(len(nama.value)))+'|')
    print('|Tempat/Tgl Lahir : '+tempat_lahir.value+', '+tanggal_lahir.value+'-'+bulan_lahir.value+'-'+tahun_lahir.value+' '*(55-int(len(tempat_lahir.value)+len(str(tanggal_lahir.value))+len(str(bulan_lahir.value))+len(str(tahun_lahir.value))))+'|' )
    print('|Jenis kelamin\t  : '+jenis_kelamin.value+' '*12+'Gol. Darah:-'+' '*(35-int(len(jenis_kelamin.value)))+'|')
    print('|Alamat\t\t  : '+alamat.value+' '*(59-int(len(alamat.value)))+'|')
    print('|    RT/RW\t  : '+rt.value+'/'+rw.value+' '*(58-int(len(rt.value)+len(rw.value)))+'|')
    print('|    Kel/Desa\t  : '+kelurahan.value+' '*(59-int(len(kelurahan.value)))+'|')
    print('|    Kecamatan\t  : '+kecamatan.value+' '*(59-int(len(kecamatan.value)))+'|')
    print('|Agama\t\t  : '+agama.value+' '*(59-int(len(agama.value)))+'|')
    print('|Status Perkawinan: '+status+' '*(59-int(len(status)))+'|')
    print('|Pekerjaan\t  : '+pekerjaan+' '*(59-int(len(pekerjaan)))+'|')
    print('|Kewarganegaraan  : '+kewarganegaraan+' '*(59-int(len(kewarganegaraan)))+'|')
    print('|Berlaku Hingga\t  : SEUMUR HIDUP'+' '*47+'|')
    print('-'*80)


def edit_data():
    file = openpyxl.load_workbook('Data_Base.xlsx')
    sheet = file.active
    while(True):
        clear_terminal()
        for i in range(2, sheet.max_row+1):
            nik = sheet.cell(column=1, row=i)
            nama = sheet.cell(column=2, row=i)
            print(i-1,f'{nik.value}', f'{nama.value}')
        baris = int(input('Masukan baris yang akan diedit : '))
        edit(sheet, baris+1)
        file.save('Data_Base.xlsx')
        ulang = input('edit data ktp lagi?(y/n)')
        if ulang == 'y':
            pass
        else:
            break
    file.close()

def edit(data,baris):
    clear_terminal()
    sheet = data
    
    nik = sheet.cell(column= 1, row=baris)
    nama = sheet.cell(column= 2, row=baris)
    tempat_lahir = sheet.cell(column= 3, row=baris)
    tanggal_lahir = sheet.cell(column= 4, row=baris)
    bulan_lahir = sheet.cell(column= 5, row=baris)
    tahun_lahir = sheet.cell(column= 6, row=baris)
    alamat = sheet.cell(column= 7, row=baris)
    rt = sheet.cell(column= 8, row=baris)
    rw = sheet.cell(column= 9, row=baris)
    kelurahan = sheet.cell(column= 10, row=baris)
    kecamatan = sheet.cell(column= 11, row=baris)
    kabupaten = sheet.cell(column= 12, row=baris)
    provinsi = sheet.cell(column= 13, row=baris)
    jenis_kelamin = sheet.cell(column= 14, row=baris)
    agama = sheet.cell(column= 15, row=baris)

    nik = nik.value
    nama = nama.value
    tempat_lahir = tempat_lahir.value
    tanggal_lahir = tanggal_lahir.value
    bulan_lahir = bulan_lahir.value
    tahun_lahir = tahun_lahir.value
    alamat = alamat.value
    rt = rt.value
    rw = rw.value
    kelurahan = kelurahan.value
    kecamatan = kecamatan.value
    kabupaten = kabupaten.value
    provinsi = provinsi.value
    jenis_kelamin = jenis_kelamin.value
    agama = agama.value

    while(True):
        clear_terminal()
        print('1. nik\t\t: ', nik)
        print('2. nama\t\t:', nama)
        print('3. tempat lahir\t:', tempat_lahir)
        print('4. tanggal lahir\t:', tanggal_lahir)
        print('5. bulan lahir\t:',bulan_lahir)
        print('6. tahun lahir\t:', tahun_lahir)
        print('7. alamat\t\t:', alamat)
        print('8. rt\t\t:', rt)
        print('9. rw\t\t:', rw)
        print('10. kelurahan\t:',kelurahan)
        print('11. kecamatan\t:',kecamatan)
        print('12. kabupaten\t:', kabupaten)
        print('13. provinsi\t:', provinsi)
        print('14. jenis_kelamin\t:', jenis_kelamin)
        print('15. agama\t\t:', agama)

        ubah = input('pilih nomor yang akan diubah : ')

        match int(ubah):
            case 1:
                nik = input('masukan nik\t\t:')
            case 2:
                nama = input('masukan nama\t\t:')
            case 3:
                tempat_lahir = input('masukan tempat lahir\t\t:')
            case 4:
                tanggal_lahir = input('masukan tanggal lahir\t\t:')
            case 5:
                bulan_lahir = input('masukan bulan lahir\t\t:')
            case 6:
                tahun_lahir = input('masukan tahun_lahir\t\t:')
            case 7:
                alamat = input('masukan alamat\t\t:')
            case 8:
                rt = input('masukan rt\t\t:')
            case 9:
                rw = input('masukan rw\t\t:')
            case 10:
                kelurahan = input('masukan kelurahan\t\t:')
            case 11:
                kecamatan = input('masukan kecamatan\t\t:')
            case 12:
                kabupaten = input('masukan kabupaten\t\t:')
            case 13:
                provinsi = input('masukan provinsi\t\t:')
            case 14:
                jenis_kelamin = input('masukan jenis kelamin\t\t:')
            case 15:
                nama = input('masukan nama\t\t:')
        ulang = input('edit data milik',nama,' lagi?(y/n)')
        if ulang == 'y':
            pass
        else:
            break
    # simpan data
    list_data = [nik, nama.upper(), tempat_lahir.upper(), tanggal_lahir, bulan_lahir, tahun_lahir, alamat.upper(), rt, rw, kelurahan.upper(), kecamatan.upper(), kabupaten.upper(), provinsi.upper(), jenis_kelamin.upper(), agama.upper()]
    kolom=0
    for i in list_data:
            kolom += 1
            sheet.cell(column=kolom, row=baris, value=i)
    clear_terminal()


def hapus_data():
    file = openpyxl.load_workbook('Data_Base.xlsx')
    sheet = file.active
    while(True):
        clear_terminal()
        for i in range(2, sheet.max_row+1):
            nik = sheet.cell(column=1, row=i)
            nama = sheet.cell(column=2, row=i)
            print(i-1,f'{nik.value}', f'{nama.value}')
        baris = int(input('Masukan baris yang akan hapus : '))
        sheet.delete_rows(idx=baris+1, amount=1)
        ulang = input('hapus ktp lagi?(y/n)')
        if ulang == 'y':
            pass
        else:
            break
    file.save('Data_Base.xlsx')
    file.close()
 

if __name__ == '__main__':
    #inisialisasi variabel
    nik = ''
    nama = ''
    tempat_lahir = ''
    tanggal_lahir = ''
    bulan_lahir = ''
    tahun_lahir = ''
    alamat = ''
    rt = ''
    rw = ''
    kelurahan = ''
    kecamatan = ''
    kabupaten = ''
    provinsi = ''
    jenis_kelamin = ''
    goldarah = '-'
    pekerjaan = 'PELAJAR'
    agama = ''
    status = 'BELUM KAWIN'
    kewarganegaraan = 'WNI'
    file = pathlib.Path('Data_Base.xlsx')
    cek_data_base(file)

    # program utama
    while(True):
        clear_terminal()
        clear_terminal()


        print('PROGRAM E-KTP')
        print('='*100)

        print('1. Tambah Data')
        print('2. Lihat Data')
        print('3. Update Data')
        print('4. Hapus Data')
        print('lainya untuk keluar')

        user_option = input('Masukan Opsi : ')

        if user_option == '1':
            tambah_data()
        elif user_option == '2':
            lihat_data()
        elif user_option == '3':
            edit_data()
        elif user_option == '4':
            hapus_data()
        else:
            clear_terminal()
            break
            # print('Delete Data')
    clear_terminal()

